#!/usr/bin/env python3
# coding:utf-8

import logging
import yaml
import click
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class Converter():

    def __init__(self):
        self._yaml_list = []
        self._header_list = [
            "No",
            "大項目",
            "中項目",
            "小項目",
            "概要",
            "手順",
            "期待値",
            "カテゴリ",
            "結果",
            "作業者",
            "作業日",
            "確認者",
            "確認日",
            "備考"]
        self._map_columns = {
            "No": 1,
            "大項目": 2,
            "中項目": 3,
            "小項目": 4,
            "description": 5,
            "procedure": 6,
            "expected_value": 7,
            "category": 8,
            "result": 9,
            "operator": 10,
            "operate_day": 11,
            "verifier": 12,
            "verify_day": 13,
            "comment": 14
            }
        self._row = 1
        self._column = 1
        self._column_number = 1
        self._num = 1

    def load_yaml(self,yaml_file):
        """yamlの読み込み"""
        with open(yaml_file) as f:
            obj = yaml.safe_load(f)
            logger.info(obj)
            self._yaml_list.append(obj)

    def _set_color(self,cell):
        #deepskyblue
        fill = PatternFill(patternType='solid', fgColor='00bfff')
        cell.fill = fill

    def _border(self, cell):
        #黒い実線
        side1 = Side(style='thin', color='000000')
        border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
        cell.border = border_aro
        
    def _create_header(self, worksheet):
        """ヘッダ行作成"""
        
        for header in self._header_list:
            cell = worksheet.cell(row=self._row,
                                  column=self._column)            
            cell.value = header
            self._border(cell)
            self._set_color(cell)
            self._column+=1

    def _fill_content(self, worksheet, obj):
        cell = worksheet.cell(row=self._row,
                              column=self._column_number)
        cell.value = self._num
        self._border(cell)        
        for key, value in obj.items():
             cell = worksheet.cell(row=self._row,
                                   column=self._map_columns[key])
             cell.value = value
             self._border(cell)             
        self._row+=1
        self._num+=1
        return True

    def _fill_suite(self, worksheet, input_list):

        for obj in input_list:
            logger.info("-----")
            logger.info(obj)
            logger.info("-----")
            if "description" in list(obj.keys()):
                self._fill_content(worksheet, obj)
            else:
                for key, value_list in obj.items():
                    cell = worksheet.cell(row=self._row,
                                          column=self._column)
                    cell.value = key
                    self._border(cell)                    
                    self._column+=1
                    self._fill_suite(worksheet, value_list)

        self._column-=1
        
    def _insert_contents(self, worksheet, obj):
        """セルに入力する"""
        for parent_suite in obj.keys():
            if parent_suite != "sheet_name":
                # 番号
                cell = worksheet.cell(row=self._row,
                                      column=self._column)
                cell.value = self._num
                self._border(cell)                
                logger.info(f"{cell},{cell.value},row={self._row},column={self._column}")
                self._column+=1
                
                # 大項目
                cell = worksheet.cell(row=self._row,
                                      column=self._column)
                cell.value = parent_suite
                self._border(cell)                
                logger.info(f"{cell},{cell.value},row={self._row},column={self._column}")                
                self._column+=1
                
                # 中項目以降
                self._fill_suite(worksheet=worksheet,
                                input_list=obj[parent_suite])
                logger.info(parent_suite)
                self._column=1

    def _export_to_excel(self, output):
        """excelに出力する"""
        logger.info(self._yaml_list)
        workbook = openpyxl.Workbook()
        for obj in self._yaml_list:
            worksheet = workbook.create_sheet(title=obj["sheet_name"])
            # ヘッダ挿入
            self._create_header(worksheet)
            self._row+=1
            self._column=1
            # 項目挿入
            self._insert_contents(worksheet, obj)
            self._row=1
            self._column=1
            self._num=1
            self.set_config(worksheet)
        workbook.save(output)
                        
    def export(self,to_format, output):
        """to_formatで指定したフォーマットでexportする"""

        if to_format == "xlsx":
            self._export_to_excel(output)


    def set_config(self,worksheet):

        # セル幅調整
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 30
        worksheet.column_dimensions['F'].width = 30
        worksheet.column_dimensions['G'].width = 30
        worksheet.column_dimensions['N'].width = 40


@click.command()
@click.option('-f','--from-format', help='source file format.', type=str, default="yaml")
@click.option('-t','--to-format', help='distination file format.', type=str, default="xlsx")
@click.option('-o','--output', help='excel file name', type=str, default="output.xlsx")
@click.argument('files', nargs=-1)
def main(from_format, to_format, output,files):
    converter = Converter()
    for f in files:
        converter.load_yaml(f)

    converter.export(to_format=to_format,
                     output=output)

if __name__ == "__main__":
    main()
